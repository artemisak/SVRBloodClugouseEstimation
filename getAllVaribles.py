import csv
import configparser
import itertools
from string import ascii_uppercase

import numpy as np
import pandas as pd
from openpyxl.reader.excel import load_workbook

config = configparser.ConfigParser()
config.read('config.ini')


def iter_all_strings():
    for size in itertools.count(1):
        for i1 in itertools.product(ascii_uppercase, repeat=size):
            yield ''.join(i1)


listFromAtoRW = []
for i in iter_all_strings():
    listFromAtoRW.append(i)
    if i == 'RW':
        break

wb = load_workbook(filename=config['MODEL']['meals_clean'])
ws = wb.active
yellow_features = []
green_features = []
blue_features = []
Y_names = []
for i in range(1, len(listFromAtoRW) + 1):
    if ws.cell(row=1, column=i).fill.fgColor.value == 'FFFFFF00':
        yellow_features.append(ws.cell(row=1, column=i).value)
    elif ws.cell(row=1, column=i).fill.fgColor.value == 'FF92D050':
        green_features.append(ws.cell(row=1, column=i).value)
    elif ws.cell(row=1, column=i).fill.fgColor.value == 'FF00B0F0':
        blue_features.append(ws.cell(row=1, column=i).value)
    elif ws.cell(row=1, column=i).fill.fgColor.value == 'FFFF0000':
        Y_names.append(ws.cell(row=1, column=i).value)

wb = load_workbook(filename=config['MODEL']['meals_param'])
ws = wb.active
mirc_features = []
for i in range(1, len(listFromAtoRW) + 1):
    if ws.cell(row=1, column=i).fill.fgColor.value == 'FFFFFF00':
        mirc_features.append(ws.cell(row=1, column=i).value)

x_values = [*yellow_features, *green_features, *blue_features, *mirc_features]
y_values = [*Y_names]
with open("X.csv", "w", newline='') as file:
    writer = csv.writer(file, delimiter=';')
    for row in x_values:
        writer.writerow([row])

with open('Y.csv', 'w', newline='') as file:
    writer = csv.writer(file, delimiter=';')
    for row in y_values:
        writer.writerow([row])

data = pd.read_excel(config["MODEL"]["meals_param"])
N = data["N"].sample(frac=1).unique()
train_num = np.random.choice(N, round(len(N) * 0.8), replace=False)
test_num = np.setdiff1d(N, train_num)

with open('train_num.csv', 'w', newline='') as file:
    for num in train_num:
        file.write(str(num) + '\n')

with open('test_num.csv', 'w', newline='') as file:
    for num in test_num:
        file.write(str(num) + '\n')
